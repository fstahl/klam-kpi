"""Smoke test for the pull_data pipeline.

Builds a minimal fake ATL workbook and a minimal kpi-template, runs
pull_data.py against them in a tmp dir, and asserts the resulting
KPIs sheet has plausible values for every declared field.

Run with:
    python3 tests/test_pull_data.py
"""
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
FIELDS_CONFIG = json.loads((ROOT / "src" / "kpi-fields.json").read_text())


def make_fake_atl(path: Path) -> None:
    """Create a minimal ATL workbook with the cells pull_data.py reads."""
    wb = Workbook()
    wb.remove(wb.active)

    params = wb.create_sheet("Parameters")
    params["C8"] = "2603"  # March 2026

    r1u = wb.create_sheet("R1U")
    r1u_q = wb.create_sheet("R1U_Q")
    r1u["C28"] = 5000;  r1u["I28"] = 50000
    r1u_q["C28"] = 15000; r1u_q["F28"] = 14000
    r1u["C51"] = 1000;  r1u["I51"] = 10000
    r1u_q["C51"] = 3000;  r1u_q["F51"] = 2800
    r1u["C52"] = 0.20;  r1u["I52"] = 0.20
    r1u_q["C52"] = 0.20;  r1u_q["F52"] = 0.20
    r1u["L87"] = 0.40
    r1u["C36"] = 3000;  r1u["I36"] = 30000
    r1u_q["C36"] = 9000;  r1u_q["F36"] = 8500
    r1u["C46"] = -1500; r1u["I46"] = -15000
    r1u_q["C46"] = -4500; r1u_q["F46"] = -4200
    r1u["C85"] = 27;    r1u["I85"] = 27
    r1u_q["C85"] = 27;    r1u_q["F85"] = 26

    link_m = wb.create_sheet("LinkM")
    # rows 57..68, columns 8 (rev), 14 (gp), 20 (admin), 27 (ebita_pct), 81 (pwc), 94 (employees)
    for i, r in enumerate(range(57, 69), start=1):
        link_m.cell(r, 8).value  = 3000 + i * 100
        link_m.cell(r, 14).value = 2000 + i * 50
        link_m.cell(r, 20).value = 800 + i * 20
        link_m.cell(r, 27).value = 0.10 + i * 0.005
        link_m.cell(r, 81).value = 0.30 + i * 0.01
        link_m.cell(r, 94).value = 25 + (i % 3)

    wb.save(path)


def make_fake_template(path: Path) -> None:
    """Create a kpi-template with the schema pull_data.py expects."""
    wb = Workbook()
    wb.remove(wb.active)

    inp = wb.create_sheet("Input")
    inp.append(["field", "Description", "MTD", "QTD", "YTD", "LASTQ"])
    # Fill with reasonable test values
    rows = [
        ("cm_count",            "", 1, 1, 3, 1),
        ("cm_value",            "", 100, 100, 300, 100),
        ("pe_count",            "", 2, 2, 6, 2),
        ("pe_value",            "", 200, 200, 600, 200),
        ("billing_target",      "", 0.85, None, None, None),
        ("pwc_target",          "", 0.55, None, None, None),
        ("converted",           "", 0, 0, 0, 0),
        ("converted_ytd",       "", 0, 0, 0, 0),
        ("converted_goal",      "", 24, None, None, None),
        ("revenue_target_fy",   "", 60000, None, None, None),
        ("ebita_target_fy",     "", 6000, None, None, None),
        ("gp_emp_target",       "", 250, None, None, None),
        ("admin_target",        "", 100, None, None, None),
        ("cm_count_target_fy",  "", 12, None, None, None),
        ("pe_count_target_fy",  "", 24, None, None, None),
    ]
    for row in rows:
        inp.append(row)

    kpis = wb.create_sheet("KPIs")
    kpis.append(["Field", "MTD", "QTD", "YTD", "LASTQ"])
    for field in FIELDS_CONFIG["fields"]:
        kpis.append([field, None, None, None, None])

    sparks = wb.create_sheet("Sparklines")
    sparks.append(["series"] + [f"M{i}" for i in range(1, 13)])

    wb.save(path)


def run_pull_data(tmpdir: Path) -> subprocess.CompletedProcess:
    return subprocess.run(
        [sys.executable, "pull_data.py"],
        cwd=tmpdir, capture_output=True, text=True,
    )


def main() -> int:
    failures = []
    with tempfile.TemporaryDirectory() as tmp_str:
        tmp = Path(tmp_str)
        # Stage everything pull_data.py needs (script, config, fake inputs)
        shutil.copy(ROOT / "pull_data.py", tmp / "pull_data.py")
        (tmp / "src").mkdir()
        shutil.copy(ROOT / "src" / "kpi-fields.json", tmp / "src" / "kpi-fields.json")
        make_fake_atl(tmp / "ATL_MB_TEST.xlsx")
        make_fake_template(tmp / "kpi-template.xlsx")

        result = run_pull_data(tmp)
        if result.returncode != 0:
            print("STDOUT:", result.stdout)
            print("STDERR:", result.stderr)
            failures.append(f"pull_data.py exited {result.returncode}")
            return _report(failures)

        wb = load_workbook(tmp / "kpi-template.xlsx", data_only=True)
        kpis = wb["KPIs"]
        rows = {kpis.cell(r, 1).value: [kpis.cell(r, c).value for c in range(2, 6)]
                for r in range(2, kpis.max_row + 1)}

        # Every config field must exist
        for field in FIELDS_CONFIG["fields"]:
            if field not in rows:
                failures.append(f"missing field row in KPIs: {field}")

        # Spot-check the MTD column
        def check(field, predicate, msg):
            v = rows.get(field, [None])[0]
            if not predicate(v):
                failures.append(f"{field}/MTD = {v!r}: {msg}")

        check("revenue",        lambda v: v == 5000,                "expected 5000 from R1U!C28")
        check("ebita",          lambda v: v == 1000,                "expected 1000 from R1U!C51")
        check("ebita_pct",      lambda v: v == 0.20,                "expected 0.20 from R1U!C52")
        check("pwc",            lambda v: v == 0.40,                "expected 0.40 from R1U!L87")
        check("pwc_target",     lambda v: v == 0.55,                "expected 0.55 propagated from Input")
        check("pwc_delta",      lambda v: abs(v - (0.40 - 0.55)) < 1e-9, "expected pwc - pwc_target")
        check("billing_target", lambda v: v == 0.85,                "expected 0.85 propagated from Input")
        check("gp_emp",         lambda v: abs(v - (3000/27)) < 1e-9, "expected GP/employees")
        check("admin",          lambda v: abs(v - (1500/27)) < 1e-9, "expected |admin| / employees")
        check("revenue_delta",  lambda v: v is not None,            "expected vs-target delta")
        check("ebita_delta",    lambda v: v is not None,            "expected vs-target delta")

        # pwc_target/billing_target propagated to all periods
        for field in ("pwc_target", "billing_target"):
            for p_idx, p_name in enumerate(["MTD","QTD","YTD","LASTQ"]):
                v = rows[field][p_idx]
                if v is None:
                    failures.append(f"{field}/{p_name} not propagated from MTD")

        # gp_emp_idx/admin_idx populated for every period when targets are set
        for field in ("gp_emp_idx", "admin_idx"):
            for p_idx, p_name in enumerate(["MTD","QTD","YTD","LASTQ"]):
                if rows[field][p_idx] is None:
                    failures.append(f"{field}/{p_name} should be populated when target is set")

    return _report(failures)


def _report(failures):
    if failures:
        print("FAIL:")
        for f in failures: print("  -", f)
        return 1
    print("OK — pull_data.py smoke test passed")
    return 0


if __name__ == "__main__":
    sys.exit(main())
