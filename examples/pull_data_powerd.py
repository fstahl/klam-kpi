# examples/pull_data_powerd.py
#
# POWER&D AB — company-specific data pipeline.
# This script is NOT part of the general Kläm KPI product.
# It exists as a reference for writing your own pull script.
#
# It reads from POWER&D's ERP exports (ATL_MB*.xlsx) and Spiris billing
# exports, computes KPIs, and writes them to kpi-template.xlsx.
#
# To write your own pipeline: read from your data source, compute your KPIs,
# and write one row per field into the KPIs sheet with columns:
#   Field | Type | MTD | QTD | YTD | LASTQ
#
# Then trigger "Data → Reload Dashboard" in the app (Cmd+Shift+R).

import json
from pathlib import Path
from openpyxl import load_workbook
from datetime import date

# =========================================================
# FILES
# =========================================================
BASE_DIR = Path(__file__).parent
ATL_FILE    = next(BASE_DIR.glob("ATL_MB*.xlsx"))
SPIRIS_FILE = next(BASE_DIR.glob("Spiris*.xlsx"), None)
KPI_FILE = BASE_DIR / "kpi-template.xlsx"
OUT_FILE = BASE_DIR / "kpi-template.xlsx"
FIELDS_CONFIG_FILE = BASE_DIR / "src" / "kpi-fields.json"

# Single source of truth for KPI field names/types
FIELDS_CONFIG = json.loads(FIELDS_CONFIG_FILE.read_text())
KNOWN_FIELDS  = set(FIELDS_CONFIG["fields"].keys())

atl_wb = load_workbook(ATL_FILE, data_only=True)
kpi_wb = load_workbook(KPI_FILE)

KPIS  = kpi_wb["KPIs"]
INPUT = kpi_wb["Input"]

# =========================================================
# SHEET REFERENCES (EN GÅNG)
# =========================================================
SHEETS = {
    "R1U": atl_wb["R1U"],
    "R1U_Q": atl_wb["R1U_Q"],
}
LINK_M = atl_wb["LinkM"]

# =========================================================
# ATL CELL-MAP VALIDATION
# =========================================================
def atl_number(sheet_key, cell, label=""):
    """Read a numeric value from an ATL cell with a clear error if it's missing/wrong type."""
    val = SHEETS[sheet_key][cell].value
    if val is None:
        raise RuntimeError(
            f"ATL cell {sheet_key}!{cell} is empty (expected number"
            + (f" for {label}" if label else "") + ")"
        )
    if not isinstance(val, (int, float)):
        raise RuntimeError(
            f"ATL cell {sheet_key}!{cell} = {val!r} ({type(val).__name__}); expected number"
            + (f" for {label}" if label else "")
        )
    return val

def linkm_number(row, col, label="", required=True):
    val = LINK_M.cell(row=row, column=col).value
    if not isinstance(val, (int, float)):
        if not required:
            return None
        raise RuntimeError(
            f"LinkM!R{row}C{col} = {val!r}; expected number"
            + (f" for {label}" if label else "")
        )
    return val

# =========================================================
# PERIOD METADATA (needed for target fractions throughout)
# =========================================================
period = str(atl_wb["Parameters"]["C8"].value)
year   = 2000 + int(period[:2])
month  = int(period[2:])
d      = date(year, month, 1)

fy_start = year if month >= 4 else year - 1
fy_month = month - 3 if month >= 4 else month + 9   # Apr=1 … Mar=12

if month in (4, 5, 6):    q = 1
elif month in (7, 8, 9):  q = 2
elif month in (10,11,12): q = 3
else:                      q = 4

q_start_fy   = (q - 1) * 3 + 1
months_in_qtd = fy_month - q_start_fy + 1

# fractions of 12 months; used to scale FY targets to each period
TARGET_FRACS = {
    "MTD":   1 / 12,
    "QTD":   months_in_qtd / 12,
    "YTD":   fy_month / 12,
    "LASTQ": 3 / 12,
}

# =========================================================
# KPI TEMPLATE HELPER
# =========================================================
def write_value(field, column_name, value):
    if field not in KNOWN_FIELDS:
        raise RuntimeError(
            f"Field '{field}' not declared in src/kpi-fields.json — add it there first"
        )
    col = next(
        (c for c in range(1, KPIS.max_column + 1)
         if KPIS.cell(1, c).value == column_name),
        None,
    )
    if col is None:
        raise RuntimeError(f"Saknar kolumn '{column_name}'")
    for r in range(2, KPIS.max_row + 1):
        if KPIS.cell(r, 1).value == field:
            KPIS.cell(r, col).value = value
            return
    raise RuntimeError(f"Saknar field '{field}'")

def read_input(field, column_name):
    """Read a value from the Input sheet by field key and period column name."""
    col = next(
        (c for c in range(1, INPUT.max_column + 1)
         if str(INPUT.cell(1, c).value or "").strip() == column_name),
        None,
    )
    if col is None:
        return None
    for r in range(2, INPUT.max_row + 1):
        if INPUT.cell(r, 1).value == field:
            return INPUT.cell(r, col).value
    return None

# =========================================================
# FASTA KÄLLOR – (ark, cell)
# =========================================================

KPI_SOURCE = {
    # -------- Revenue --------
    ("revenue", "MTD"):   ("R1U",   "C28"),
    ("revenue", "QTD"):   ("R1U_Q", "C28"),
    ("revenue", "YTD"):   ("R1U",   "I28"),
    ("revenue", "LASTQ"): ("R1U_Q", "F28"),


    # -------- EBITA --------
    ("ebita", "MTD"):   ("R1U",   "C51"),
    ("ebita", "QTD"):   ("R1U_Q", "C51"),
    ("ebita", "YTD"):   ("R1U",   "I51"),
    ("ebita", "LASTQ"): ("R1U_Q", "F51"),


    # -------- EBITA % = ROS EBITA --------
    ("ebita_pct", "MTD"):   ("R1U",   "C52"),   # ROS EBITA MTD
    ("ebita_pct", "QTD"):   ("R1U_Q", "C52"),   # ROS EBITA QTD
    ("ebita_pct", "YTD"):   ("R1U",   "I52"),   # ROS EBITA YTD
    ("ebita_pct", "LASTQ"): ("R1U_Q", "F52"),   # ROS EBITA LASTQ

    # -------- REX (EBITA) --------
    ("rex", "MTD"):   ("R1U",   "C96"),
    ("rex", "QTD"):   ("R1U_Q", "C94"),
    ("rex", "YTD"):   ("R1U",   "I96"),
    ("rex", "LASTQ"): ("R1U_Q", "F94"),

    # -------- PWC (rullande 12M) --------
    ("pwc", "MTD"):   ("R1U", "L87"),
    ("pwc", "QTD"):   ("R1U", "L87"),
    ("pwc", "YTD"):   ("R1U", "L87"),
    ("pwc", "LASTQ"): ("R1U", "L87"),
}

# =========================================================
# POPULATE KPI VALUES (GENERISKT)
# =========================================================
for (field, _p), (sheet_key, cell) in KPI_SOURCE.items():
    write_value(field, _p, atl_number(sheet_key, cell, label=f"{field}/{_p}"))

# =========================================================
# MANUAL INPUT — read from Input sheet, write to KPIs
# =========================================================
# Per-period fields: read each period column independently
MANUAL_FIELDS = [
    "cm_count", "cm_value",
    "pe_count", "pe_value",
    "converted", "converted_ytd",
    "converted_goal",
]
for field in MANUAL_FIELDS:
    for _p in ["MTD", "QTD", "YTD", "LASTQ"]:
        value = read_input(field, _p)
        if value is not None:
            write_value(field, _p, value)

# Single-value fields: stored once in MTD column, propagated to all periods
SINGLE_VALUE_FIELDS = ["pwc_target", "billing_target"]
for field in SINGLE_VALUE_FIELDS:
    value = read_input(field, "MTD")
    if value is not None:
        for _p in ["MTD", "QTD", "YTD", "LASTQ"]:
            write_value(field, _p, value)

# =========================================================
# COMPUTED: GP per employee (gross profit / avg employees)
# =========================================================
GP_SOURCES = {
    "MTD":   ("R1U",   "C36", "C85"),
    "QTD":   ("R1U_Q", "C36", "C85"),
    "YTD":   ("R1U",   "I36", "I85"),
    "LASTQ": ("R1U_Q", "F36", "F85"),
}
gp_emp_values = {}
for _p, (sheet_key, gp_cell, emp_cell) in GP_SOURCES.items():
    gp = atl_number(sheet_key, gp_cell, label=f"gp/{_p}")
    emp = atl_number(sheet_key, emp_cell, label=f"employees/{_p}")
    gp_emp_values[_p] = gp / emp
    write_value("gp_emp", _p, gp_emp_values[_p])

period_months = {"MTD": 1, "QTD": months_in_qtd, "YTD": fy_month, "LASTQ": 3}

gp_emp_target = read_input("gp_emp_target", "MTD")
for _p, _v in gp_emp_values.items():
    if gp_emp_target:
        pt = gp_emp_target * period_months[_p]
        write_value("gp_emp_delta", _p, (_v - pt) / pt)
    else:
        write_value("gp_emp_delta", _p, None)

# gp_emp_idx: actual vs period target (100 = on target)
for _p, _v in gp_emp_values.items():
    if gp_emp_target:
        pt = gp_emp_target * period_months[_p]
        write_value("gp_emp_idx", _p, (_v / pt) * 100)
    else:
        write_value("gp_emp_idx", _p, None)

# =========================================================
# COMPUTED: Admin cost per employee (overhead / avg employees)
# =========================================================
ADMIN_SOURCES = {
    "MTD":   ("R1U",   "C46", "C85"),
    "QTD":   ("R1U_Q", "C46", "C85"),
    "YTD":   ("R1U",   "I46", "I85"),
    "LASTQ": ("R1U_Q", "F46", "F85"),
}

def calc_admin(sheet_key, expense_cell, emp_cell, label=""):
    return abs(atl_number(sheet_key, expense_cell, label)) / atl_number(sheet_key, emp_cell, label)

admin_values = {
    "MTD":   calc_admin("R1U",   "C46", "C85", "admin/MTD"),
    "QTD":   calc_admin("R1U_Q", "C46", "C85", "admin/QTD"),
    "YTD":   calc_admin("R1U",   "I46", "I85", "admin/YTD"),
    "LASTQ": calc_admin("R1U_Q", "F46", "F85", "admin/LASTQ"),
}
for _p, _v in admin_values.items():
    write_value("admin", _p, _v)

# =========================================================
# COMPUTED: pwc_delta = actual P/WC minus target (pp difference)
# =========================================================
pwc_val        = atl_number("R1U", "L87", "pwc")  # rolling 12M, same for all periods
pwc_target_val = read_input("pwc_target", "MTD")
for _p in ["MTD", "QTD", "YTD", "LASTQ"]:
    if pwc_target_val:
        write_value("pwc_delta", _p, pwc_val - pwc_target_val)
    else:
        write_value("pwc_delta", _p, None)

# =========================================================
# COMPUTED: admin_delta and admin_idx
# =========================================================
def read_value(field, column_name):
    col = next((c for c in range(1, KPIS.max_column + 1) if KPIS.cell(1, c).value == column_name), None)
    if col is None:
        return None
    for r in range(2, KPIS.max_row + 1):
        if KPIS.cell(r, 1).value == field:
            return KPIS.cell(r, col).value
    return None

# admin_delta: % above/below target (positive = over budget, shown inverted in UI)
admin_target = read_input("admin_target", "MTD")
for _p, _v in admin_values.items():
    if admin_target:
        pt = admin_target * period_months[_p]
        write_value("admin_delta", _p, (_v - pt) / pt)
    else:
        write_value("admin_delta", _p, None)

# admin_idx: actual vs period target (100 = on target)
for _p, _v in admin_values.items():
    if admin_target:
        pt = admin_target * period_months[_p]
        write_value("admin_idx", _p, (_v / pt) * 100)
    else:
        write_value("admin_idx", _p, None)

# =========================================================
# COMPUTED: vs-target deltas (revenue, ebita, cm, pe)
# =========================================================
_period_months = {"MTD": 1, "QTD": months_in_qtd, "YTD": fy_month, "LASTQ": 3}

def _write_fy_target_delta(actual_field, delta_field, target_fy):
    for _p in ["MTD", "QTD", "YTD", "LASTQ"]:
        actual = read_value(actual_field, _p)
        if target_fy and actual is not None:
            pt = target_fy * (_period_months[_p] / 12)
            write_value(delta_field, _p, (actual - pt) / pt if pt else None)
        else:
            write_value(delta_field, _p, None)

_write_fy_target_delta("revenue",  "revenue_delta", read_input("revenue_target_fy", "MTD"))
_write_fy_target_delta("ebita",    "ebita_delta",   read_input("ebita_target_fy", "MTD"))
_write_fy_target_delta("cm_count", "cm_delta",      read_input("cm_count_target_fy", "MTD"))
_write_fy_target_delta("pe_count", "pe_delta",      read_input("pe_count_target_fy", "MTD"))

# =========================================================
# LABEL & RANGE (SAMMA SOM TIDIGARE)
# =========================================================
fy_label = f"FY{str(fy_start)[-2:]}/{str(fy_start+1)[-2:]}"

write_value("label", "MTD", d.strftime("%b %Y"))
write_value("label", "QTD", f"Last 3M")
write_value("label", "YTD", f"YTD {fy_label}")

prev_q = 4 if q == 1 else q - 1
prev_fy = fy_start - 1 if q == 1 else fy_start
write_value("label", "LASTQ", f"Q{prev_q} FY{str(prev_fy)[-2:]}/{str(prev_fy+1)[-2:]}")

write_value("range", "MTD", d.strftime("%b %Y"))
write_value("range", "YTD", f"Apr {fy_start} – {d.strftime('%b %Y')}")

if q == 1: qs = date(year,4,1)
elif q == 2: qs = date(year,7,1)
elif q == 3: qs = date(year,10,1)
else: qs = date(year,1,1)

write_value("range", "QTD", f"{qs.strftime('%b')}–{d.strftime('%b %Y')}")

# LASTQ range
if prev_q == 1: lqs, lqe = date(year-1,4,1), date(year-1,6,30)
elif prev_q == 2: lqs, lqe = date(year-1,7,1), date(year-1,9,30)
elif prev_q == 3: lqs, lqe = date(year-1,10,1), date(year-1,12,31)
else: lqs, lqe = date(year,1,1), date(year,3,31)

write_value("range", "LASTQ", f"{lqs.strftime('%b')}–{lqe.strftime('%b %Y')}")

# =========================================================
# SPARKLINES – LinkM rows 57-68 (oldest → latest)
# =========================================================
SPARKS_SHEET = kpi_wb["Sparklines"]
SPARK_ROW_RANGE = range(57, 69)  # 12 months

SPARK_SOURCES = {
    "rev":       8,   # column H
    "ebita_pct": 27,  # column AA
    "pwc":       81,  # column CC
}

def write_sparkline(key, values):
    target_row = None
    for r in range(2, SPARKS_SHEET.max_row + 1):
        if SPARKS_SHEET.cell(r, 1).value == key:
            target_row = r
            break
    if target_row is None:
        target_row = SPARKS_SHEET.max_row + 1
        SPARKS_SHEET.cell(target_row, 1).value = key
    for i, v in enumerate(values):
        SPARKS_SHEET.cell(target_row, i + 2).value = v
    # Clear any trailing cells from previous longer writes
    for c in range(len(values) + 2, SPARKS_SHEET.max_column + 1):
        SPARKS_SHEET.cell(target_row, c).value = None

for key, col in SPARK_SOURCES.items():
    values = [linkm_number(r, col, f"{key} sparkline", required=False) for r in SPARK_ROW_RANGE]
    write_sparkline(key, values)

# Computed sparklines: GP and admin per employee (monthly)
def _spark_div(num, den):
    return (num / den) if (num is not None and den) else None

gp_per_emp  = [_spark_div(linkm_number(r, 14, required=False), linkm_number(r, 94, required=False)) for r in SPARK_ROW_RANGE]
adm_per_emp = [_spark_div(linkm_number(r, 20, required=False), linkm_number(r, 94, required=False)) for r in SPARK_ROW_RANGE]
write_sparkline("gp",  gp_per_emp)
write_sparkline("adm", adm_per_emp)

# =========================================================
# BILLING RATE — from Spiris export (last row of column F)
# =========================================================
if SPIRIS_FILE:
    spiris_wb = load_workbook(SPIRIS_FILE, data_only=True)
    spiris_sheet = spiris_wb.active
    last_row = max(
        r for r in range(1, spiris_sheet.max_row + 1)
        if spiris_sheet.cell(r, 6).value is not None
    )
    billing_rate = spiris_sheet.cell(last_row, 6).value / 100

    # ── Rolling 12-month billing history ──────────────────
    if "BillingHistory" not in kpi_wb.sheetnames:
        hist = kpi_wb.create_sheet("BillingHistory")
        hist.append(["Date", "Billing"])
    else:
        hist = kpi_wb["BillingHistory"]

    # Billing month matches the ATL reporting period
    billing_key = d.isoformat()

    # Read all rows, overwrite-by-date or append, re-sort, keep 12 most recent
    hist_data = {
        row[0]: row[1]
        for row in hist.iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    hist_data[billing_key] = billing_rate
    sorted_items = sorted(hist_data.items(), key=lambda kv: kv[0])[-12:]

    # Clear existing data rows, then rewrite in order
    if hist.max_row >= 2:
        hist.delete_rows(2, hist.max_row - 1)
    for date_iso, rate in sorted_items:
        hist.append([date_iso, rate])

    # Compute per-period billing averages from history
    fy_start_iso = f"{fy_start}-04-01"
    hist_rows = [(row[0], row[1]) for row in hist.iter_rows(min_row=2, values_only=True) if row[0]]

    def avg(rates):
        return sum(rates) / len(rates) if rates else None

    ytd_rates   = [v for k, v in hist_rows if k >= fy_start_iso]
    qtd_rates   = [v for _, v in hist_rows[-3:]]           # last 3 months
    lastq_rates = [v for _, v in hist_rows[-6:-3]]         # 3 months before that

    billing_by_period = {
        "MTD":   billing_rate,
        "QTD":   avg(qtd_rates),
        "YTD":   avg(ytd_rates),
        "LASTQ": avg(lastq_rates),
    }

    billing_target_val = read_value("billing_target", "MTD")
    for _p, _v in billing_by_period.items():
        write_value("billing", _p, _v)
        write_value(
            "billing_delta", _p,
            (_v - billing_target_val) if (_v is not None and billing_target_val is not None) else None,
        )

    # Update billing sparkline from full history (oldest → latest)
    write_sparkline("bill", [r[1] for r in hist.iter_rows(min_row=2, values_only=True)])
else:
    print("Varning: Ingen Spiris-fil hittades, billing rate ej uppdaterad")

# =========================================================
# SAVE
# =========================================================
kpi_wb.save(OUT_FILE)
print("KLART – script med tydliga ark- och cellreferenser")
